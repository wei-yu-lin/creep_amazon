import json
import requests
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import pandas
import time
import os
from openpyxl import load_workbook
from shutil import move
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import WebDriverWait


url = "https://www.amazon.com/-/zh_TW/s?i=electronics-intl-ship&bbn=16225009011&rh=n%3A541966%2Cn%3A193870011%2Cn%3A17923671011%2Cn%3A284822&dc&page={}&language=zh_TW&qid=1598425548&rnid=17923671011&ref=sr_pg_{}"
options = Options()
options.add_argument("--disable-notifications")
options.add_argument('headless')
chrome = webdriver.Chrome('./chromedriver', chrome_options=options)
amazon_allpage = []
data = []
tStart = time.time()
for i in range(0, 2):
    chrome.get(url.format(i, i))
    soup = BeautifulSoup(chrome.page_source, 'lxml')
    amazon_card = soup.find_all(
        class_="a-size-medium a-color-base a-text-normal")
    data.append([ele.text for ele in amazon_card])
#     amazon_allpage += [amazon_card]
chrome.close()
tEnd = time.time()
print("�`�@�Ӯ�:%f ��" % (tEnd-tStart))

amazon_json = json.load(
    open("AMAZON_COMPUTER_PRICE" + ".json", encoding="utf-8"))
df = pd.DataFrame(data[1:], columns=data[0])


now_date = datetime.strftime(datetime.now(), '%Y%m%d')
# �S�ɮת��ܫh�إ�
output_file = now_date + "_amazon_graphics card.xlsx"
target_path = "D:\\Amazon_graphics_card\\"
if os.path.exists(output_file) == False:
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    writer.book.create_sheet('default')  # �s�W��sheet
else:  # �p�G�ɮצs�b�Aappend
    book = load_workbook(output_file)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    writer.book = book

if amazon_json['sheet_name'] in writer.book.sheetnames:
    writer.book.remove(writer.book[amazon_json['sheet_name']])

    # �P�_AMAZON���άO�_���\
# if len(df) > 0:
#     # ���X��sheet�һݭn����ltable���
#     df_tmp = df.loc[:, amazon_json['raw_table_cols']]
#     # ���user�Ϊ�column name
#     df_tmp.columns = sheet_dict['excel_cols_title']
# else:
#     df_tmp = pd.DataFrame(columns=sheet_dict['excel_cols_title'])
#     df_tmp = df_tmp.append({'���F�s��': '����L�ƾ�'}, ignore_index=True)

# �g�Jexcel sheet
df.to_excel(writer, sheet_name=amazon_json['sheet_name'], index=False)
if (len(writer.book.sheetnames) > 1):  # �Y����Lsheet�A�R��default sheet
    if 'default' in writer.book.sheetnames:
        writer.book.remove(writer.book['default'])
writer.save()


cwd = os.getcwd()
curr_file = cwd + "\\" + output_file
target_file = target_path + output_file

try:
    move(curr_file, target_file)
    print(output_file + "\n�����ɮק���.")
except Exception as e:
    print(output_file + "\n�����ɮץ���.")
    print(e.args)
